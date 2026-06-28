import csv
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions
from django.db.models import Sum
from datetime import datetime

# ReportLab libraries for PDF export
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

# Openpyxl for Excel export
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Models
from apps.cars.models import Car
from apps.expenses.models import Expense
from apps.earnings.models import EarningUpload
from apps.settlements.models import Settlement
from apps.accounts.permissions import IsManagerOrAbove


class VehicleProfitabilityReportView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsManagerOrAbove]

    def get_profitability_data(self):
        cars = Car.objects.all()
        report_data = []

        for car in cars:
            # Sum all expenses directly assigned to this vehicle
            total_expenses = Expense.objects.filter(car=car).aggregate(total=Sum('amount'))['total'] or 0.00
            
            # Sum all earnings of the driver currently assigned to this vehicle (mock representing that vehicle's share)
            total_earnings = 0.00
            if car.driver_assignment:
                total_earnings = EarningUpload.objects.filter(
                    driver=car.driver_assignment,
                    ocr_status=EarningUpload.OCRStatuses.SUCCESS
                ).aggregate(total=Sum('gross_earnings'))['total'] or 0.00
                
            net_profit = total_earnings - float(total_expenses)
            
            report_data.append({
                "registration_number": car.registration_number,
                "brand": car.brand,
                "model": car.model,
                "assigned_driver": car.driver_assignment.name if car.driver_assignment else "None",
                "total_earnings": float(total_earnings),
                "total_expenses": float(total_expenses),
                "net_profit": float(net_profit),
                "status": car.status
            })
            
        return report_data

    def get(self, request):
        export_format = request.query_params.get('format', 'json').lower()
        data = self.get_profitability_data()

        if export_format == 'json':
            return Response(data)
            
        elif export_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="vehicle_profitability.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Reg Number', 'Brand', 'Model', 'Driver', 'Earnings', 'Expenses', 'Net Profit', 'Status'])
            
            for row in data:
                writer.writerow([
                    row['registration_number'], row['brand'], row['model'],
                    row['assigned_driver'], row['total_earnings'], row['total_expenses'],
                    row['net_profit'], row['status']
                ])
            return response
            
        elif export_format == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Vehicle Profitability"
            
            # Formatting
            title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            
            ws.merge_cells('A1:H1')
            ws['A1'] = "Vehicle Profitability Report"
            ws['A1'].font = title_font
            ws['A1'].fill = title_fill
            ws['A1'].alignment = Alignment(horizontal="center")
            
            headers = ['Reg Number', 'Brand', 'Model', 'Driver', 'Total Earnings (₹)', 'Total Expenses (₹)', 'Net Profit (₹)', 'Status']
            ws.append([])  # blank row
            ws.append(headers)
            
            # Apply header fonts
            for col_idx in range(1, 9):
                cell = ws.cell(row=3, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                
            for item in data:
                ws.append([
                    item['registration_number'], item['brand'], item['model'],
                    item['assigned_driver'], item['total_earnings'], item['total_expenses'],
                    item['net_profit'], item['status']
                ])
                
            # Formatting numbers
            for row in range(4, len(data) + 4):
                ws.cell(row=row, column=5).number_format = '₹#,##0.00'
                ws.cell(row=row, column=6).number_format = '₹#,##0.00'
                ws.cell(row=row, column=7).number_format = '₹#,##0.00'

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="vehicle_profitability.xlsx"'
            wb.save(response)
            return response
            
        elif export_format == 'pdf':
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="vehicle_profitability.pdf"'
            
            doc = SimpleDocTemplate(response, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            story.append(Paragraph("Fleet Vehicle Profitability Report", styles['Title']))
            story.append(Spacer(1, 20))
            
            # Table data
            table_data = [['Reg Number', 'Driver', 'Earnings', 'Expenses', 'Net Profit']]
            for item in data:
                table_data.append([
                    item['registration_number'],
                    item['assigned_driver'],
                    f"INR {item['total_earnings']:.2f}",
                    f"INR {item['total_expenses']:.2f}",
                    f"INR {item['net_profit']:.2f}"
                ])
                
            t = Table(table_data)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.navy),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(t)
            doc.build(story)
            return response
            
        return Response({"error": "Unsupported format"}, status=400)


class SettlementReportView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsManagerOrAbove]

    def get(self, request):
        export_format = request.query_params.get('format', 'json').lower()
        settlements = Settlement.objects.all().order_by('-start_date')
        
        if export_format == 'json':
            data = []
            for s in settlements:
                data.append({
                    "driver": s.driver.name,
                    "period": f"{s.start_date} to {s.end_date}",
                    "gross": float(s.gross_earnings),
                    "cash": float(s.cash_collected),
                    "expenses": float(s.expenses_amount),
                    "advances": float(s.advances_amount),
                    "commission": float(s.commission_amount),
                    "final_payout": float(s.final_settlement_amount),
                    "status": s.status
                })
            return Response(data)
            
        elif export_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="settlements_report.csv"'
            writer = csv.writer(response)
            writer.writerow(['Driver', 'Start Date', 'End Date', 'Gross', 'Cash Collected', 'Expenses', 'Advances', 'Commission', 'Final Payout', 'Status'])
            
            for s in settlements:
                writer.writerow([
                    s.driver.name, s.start_date, s.end_date, s.gross_earnings,
                    s.cash_collected, s.expenses_amount, s.advances_amount,
                    s.commission_amount, s.final_settlement_amount, s.status
                ])
            return response
            
        return Response({"detail": "Format only support json/csv for this report."}, status=400)
